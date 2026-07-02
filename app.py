import base64
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

import pandas as pd
import streamlit as st
import xlwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import create_client


st.set_page_config(
    page_title="ITSSystems Cost Calculator",
    page_icon="€",
    layout="wide",
)

st.markdown(
    """
    <style>
    .st-key-save_part_action button {
        background-color: #198754 !important;
        border-color: #198754 !important;
        color: white !important;
        font-weight: 700 !important;
    }
    .st-key-update_part_action button {
        background-color: #dc3545 !important;
        border-color: #dc3545 !important;
        color: white !important;
        font-weight: 700 !important;
    }
    .st-key-save_part_action button:hover {
        background-color: #157347 !important;
        border-color: #146c43 !important;
    }
    .st-key-update_part_action button:hover {
        background-color: #bb2d3b !important;
        border-color: #b02a37 !important;
    }
    .st-key-unavailable_materials_area {
        background-color: rgba(220, 53, 69, 0.13);
        border: 1px solid #dc3545;
        border-radius: 12px;
        padding: 18px;
        margin-top: 18px;
    }
    .st-key-unavailable_materials_area button {
        background-color: #dc3545 !important;
        border-color: #dc3545 !important;
        color: white !important;
        font-weight: 700 !important;
    }
    .st-key-unavailable_materials_area button:hover {
        background-color: #bb2d3b !important;
        border-color: #b02a37 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# GİRİŞ VE VERİTABANI
# =========================================================
def check_password():
    if st.session_state.get("authenticated"):
        return True

    st.title("ITSSystems Cost Calculator")
    st.caption("Devam etmek için uygulama şifresini girin.")

    password = st.text_input("Şifre", type="password")
    if st.button("Giriş yap", type="primary"):
        if password == st.secrets["APP_PASSWORD"]:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Şifre hatalı.")

    return False


if not check_password():
    st.stop()


@st.cache_resource
def get_db():
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_SERVICE_ROLE_KEY"],
    )


db = get_db()


# =========================================================
# GENEL YARDIMCI FONKSİYONLAR
# =========================================================
def parse_decimal(value, default=None):
    if value is None:
        return default

    text = str(value).strip()
    if not text:
        return default

    text = (
        text.replace("€", "")
        .replace("₺", "")
        .replace("TL", "")
        .replace(" ", "")
    )

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return default


def format_number(value, digits=2):
    return (
        f"{float(value):,.{digits}f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def format_eur(value):
    return f"{format_number(value)} €"


def format_tl(value):
    return f"{format_number(value)} TL"


def convert_price(source_value, source_currency, eur_tl_rate):
    source_value = float(source_value or 0)
    eur_tl_rate = float(eur_tl_rate or 1)

    if source_currency == "TL":
        tl_value = source_value
        eur_value = source_value / eur_tl_rate if eur_tl_rate else 0
    else:
        eur_value = source_value
        tl_value = source_value * eur_tl_rate

    return eur_value, tl_value


def calculate_rectangular_material(
    length_mm,
    width_mm,
    height_mm,
    density_g_cm3,
):
    volume_mm3 = (
        float(length_mm)
        * float(width_mm)
        * float(height_mm)
    )
    volume_cm3 = volume_mm3 / 1000
    weight_kg = (
        volume_mm3
        * float(density_g_cm3)
        / 1_000_000
    )
    return volume_mm3, volume_cm3, weight_kg


def get_settings():
    response = db.table("ayarlar").select("*").eq("id", 1).execute()

    if response.data:
        return response.data[0]

    defaults = {
        "id": 1,
        "eur_tl_kuru": 50,
        "logo_base64": None,
    }
    db.table("ayarlar").insert(defaults).execute()
    return defaults


def update_settings(rate=None, logo_base64="__KEEP__"):
    current = get_settings()

    payload = {
        "id": 1,
        "eur_tl_kuru": (
            float(rate)
            if rate is not None
            else float(current.get("eur_tl_kuru") or 50)
        ),
        "logo_base64": (
            current.get("logo_base64")
            if logo_base64 == "__KEEP__"
            else logo_base64
        ),
    }

    db.table("ayarlar").upsert(payload).execute()


def get_prices():
    response = (
        db.table("fiyat_tanimlari")
        .select("*")
        .order("ad")
        .execute()
    )
    return response.data or []


def get_labors():
    response = (
        db.table("iscilik_tanimlari")
        .select("*")
        .order("ad")
        .execute()
    )
    return response.data or []


def get_parts():
    response = (
        db.table("parcalar")
        .select("*")
        .order("created_at", desc=True)
        .execute()
    )
    return response.data or []


def get_part_items(part_id=None):
    query = db.table("parca_kalemleri").select("*")

    if part_id is not None:
        query = query.eq("parca_id", part_id)

    return query.execute().data or []


def get_part_labors(part_id=None):
    query = db.table("parca_iscilik_kalemleri").select("*")

    if part_id is not None:
        query = query.eq("parca_id", part_id)

    return query.execute().data or []



def get_supplier_materials():
    response = (
        db.table("tedarikci_malzemeleri")
        .select("*")
        .order("malzeme_adi")
        .order("tedarikci_adi")
        .execute()
    )
    return response.data or []


def get_unavailable_materials():
    response = (
        db.table("bulunamayan_malzemeler")
        .select("*")
        .order("malzeme_adi")
        .execute()
    )
    return response.data or []


def normalize_web_url(value):
    url = (value or "").strip()

    if not url:
        return ""

    if not url.lower().startswith(("http://", "https://")):
        url = f"https://{url}"

    return url


def get_price_source(item):
    currency = item.get("kaynak_para_birimi") or "EUR"
    source_value = item.get("kaynak_birim_fiyat")

    if source_value is None:
        source_value = item.get("birim_fiyat_eur", 0)

    return currency, float(source_value or 0)


def get_labor_source(item):
    currency = item.get("kaynak_para_birimi") or "EUR"
    source_value = item.get("kaynak_saatlik_ucret") or 0
    return currency, float(source_value)


def get_density(item):
    value = item.get("yogunluk_g_cm3")
    return float(value) if value is not None else 0.0


# =========================================================
# BAŞLIK, LOGO VE KUR
# =========================================================
def render_header(settings):
    logo_col, info_col = st.columns([1, 4], vertical_alignment="center")

    logo_value = settings.get("logo_base64")

    with logo_col:
        if logo_value:
            try:
                logo_bytes = base64.b64decode(logo_value)
                st.image(io.BytesIO(logo_bytes), width=170)
            except Exception:
                st.warning("Logo görüntülenemedi.")

        popover_title = "Logo ayarları" if logo_value else "Logo yükle"

        with st.popover(popover_title, use_container_width=True):
            st.caption(
                "Önerilen logo: kare formatta şeffaf PNG, "
                "1024 × 1024 px ve 2 MB'tan küçük."
            )

            logo_action = st.radio(
                "İşlem",
                ["Değiştir", "Kaldır"] if logo_value else ["Yükle"],
                horizontal=True,
                key="logo_action",
            )

            if logo_action in {"Yükle", "Değiştir"}:
                logo_file = st.file_uploader(
                    "Kare PNG veya JPEG seç",
                    type=["png", "jpg", "jpeg"],
                    key="logo_file",
                )

                if st.button(
                    "Logoyu kaydet",
                    key="save_logo",
                    use_container_width=True,
                ):
                    if logo_file is None:
                        st.warning("Önce bir logo dosyası seç.")
                    elif logo_file.size > 2_000_000:
                        st.error("Logo dosyası 2 MB'tan küçük olmalı.")
                    else:
                        encoded = base64.b64encode(
                            logo_file.getvalue()
                        ).decode("utf-8")
                        update_settings(logo_base64=encoded)
                        st.success("Logo kaydedildi.")
                        st.rerun()

            elif logo_action == "Kaldır":
                st.warning("Logo kalıcı olarak kaldırılacak.")
                if st.button(
                    "Kaldırmayı onayla",
                    key="remove_logo",
                    use_container_width=True,
                ):
                    update_settings(logo_base64=None)
                    st.rerun()

    with info_col:
        st.markdown(
            """
            <div style="font-size:1.05rem; line-height:1.55; color:#b9bec8; max-width:900px;">
            Parça bazında malzeme, kaplama, ölçüm, ek işlem ve işçilik maliyetlerini
            hesaplayın; sonuçları EUR ve TL olarak kaydedin, güncelleyin ve Excel'e aktarın.
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_rate(settings):
    current_rate = float(settings.get("eur_tl_kuru") or 50)

    with st.expander("EUR / TL Kur Ayarı", expanded=False):
        with st.form("exchange_rate_form"):
            col1, col2, col3 = st.columns([1.3, 1, 3])

            with col1:
                rate_text = st.text_input(
                    "1 EUR kaç TL?",
                    value=format_number(current_rate, 4),
                )

            with col2:
                st.write("")
                st.write("")
                save_rate = st.form_submit_button(
                    "Kuru kaydet",
                    type="primary",
                    use_container_width=True,
                )

            with col3:
                st.info(
                    "TL girilen fiyatın TL değeri sabit kalır; EUR karşılığı kura göre değişir. "
                    "EUR girilen fiyatın EUR değeri sabit kalır; TL karşılığı kura göre değişir."
                )

            if save_rate:
                new_rate = parse_decimal(rate_text)

                if new_rate is None or new_rate <= 0:
                    st.error("Geçerli ve sıfırdan büyük bir kur gir.")
                else:
                    update_settings(rate=new_rate)
                    st.success("Kur güncellendi.")
                    st.rerun()


# =========================================================
# XLS DIŞA AKTARMA
# =========================================================
def build_export_frames(rate):
    prices = get_prices()
    labors = get_labors()
    parts = get_parts()
    part_items = get_part_items()
    part_labors = get_part_labors()

    price_map = {item["id"]: item for item in prices}
    labor_map = {item["id"]: item for item in labors}

    summary_rows = []
    detail_rows = []

    for part in parts:
        part_id = part["id"]
        quantity = int(part["adet"])
        single_eur = 0.0
        single_tl = 0.0
        operations = []

        for row in [x for x in part_items if x["parca_id"] == part_id]:
            definition = price_map.get(row["fiyat_tanimi_id"], {})
            currency = row.get("kaynak_para_birimi") or "EUR"
            source_value = row.get("kaynak_birim_fiyat")
            if source_value is None:
                source_value = row.get("birim_fiyat_eur", 0)

            unit_eur, unit_tl = convert_price(
                float(source_value or 0), currency, rate
            )
            amount = float(row.get("miktar") or 0)
            amount_type = row.get("miktar_turu") or "adet"
            line_eur = unit_eur * amount
            line_tl = unit_tl * amount
            single_eur += line_eur
            single_tl += line_tl

            if amount_type == "kg":
                amount_text = f"{format_number(amount, 4)} kg"
            elif amount_type == "saat":
                amount_text = f"{format_number(amount, 4)} saat"
            else:
                amount_text = f"{format_number(amount, 0)} adet"
            operations.append(
                f'{definition.get("ad", "")} [{amount_text}] '
                f'({format_eur(line_eur)} / {format_tl(line_tl)})'
            )
            detail_rows.append({
                "Parça Adı": part["parca_adi"],
                "Kategori": definition.get("kategori", ""),
                "İşlem / Kalem": definition.get("ad", ""),
                "Miktar": amount,
                "Miktar Türü": amount_type,
                "Tek Parça Kalem EUR": line_eur,
                "Tek Parça Kalem TL": line_tl,
            })

        for row in [x for x in part_labors if x["parca_id"] == part_id]:
            definition = labor_map.get(row["iscilik_tanimi_id"], {})
            currency = row.get("kaynak_para_birimi") or "EUR"
            source_value = float(row.get("kaynak_saatlik_ucret") or 0)
            hourly_eur, hourly_tl = convert_price(source_value, currency, rate)
            hours = float(row.get("saat") or 0)
            line_eur = hourly_eur * hours
            line_tl = hourly_tl * hours
            single_eur += line_eur
            single_tl += line_tl

            operations.append(
                f'{definition.get("ad", "")} [{format_number(hours, 4)} saat] '
                f'({format_eur(line_eur)} / {format_tl(line_tl)})'
            )
            detail_rows.append({
                "Parça Adı": part["parca_adi"],
                "Kategori": "İşçilik",
                "İşlem / Kalem": definition.get("ad", ""),
                "Miktar": hours,
                "Miktar Türü": "saat",
                "Tek Parça Kalem EUR": line_eur,
                "Tek Parça Kalem TL": line_tl,
            })

        dimensions = ""
        if all(part.get(key) is not None for key in ("boy_mm", "en_mm", "yukseklik_mm")):
            dimensions = (
                f'{format_number(part["boy_mm"], 2)} × '
                f'{format_number(part["en_mm"], 2)} × '
                f'{format_number(part["yukseklik_mm"], 2)} mm'
            )

        summary_rows.append({
            "Parça Adı": part["parca_adi"],
            "Adet": quantity,
            "Kaba Ebat": dimensions,
            "Malzeme Ağırlığı (kg)": float(part.get("malzeme_agirlik_kg") or 0),
            "İşlemler": " + ".join(operations),
            "Birim Fiyat EUR": single_eur,
            "Birim Fiyat TL": single_tl,
            "Toplam Fiyat EUR": single_eur * quantity,
            "Toplam Fiyat TL": single_tl * quantity,
            "Kullanılan EUR/TL Kuru": rate,
        })

    return {
        "Liste": pd.DataFrame(summary_rows),
        "Detaylar": pd.DataFrame(detail_rows),
    }


def build_xls(rate):
    frames = build_export_frames(rate)
    workbook = xlwt.Workbook(encoding="utf-8")
    header_style = xlwt.easyxf(
        "font: bold on; pattern: pattern solid, fore_colour gray25;"
    )
    text_style = xlwt.easyxf()
    number_style = xlwt.easyxf(num_format_str="# ,##0.0000".replace(" ", ""))
    money_style = xlwt.easyxf(num_format_str="# ,##0.00".replace(" ", ""))

    for sheet_name, dataframe in frames.items():
        sheet = workbook.add_sheet(sheet_name[:31])
        if dataframe.empty:
            sheet.write(0, 0, "Kayıt bulunmuyor.", text_style)
            continue

        for column_index, column_name in enumerate(dataframe.columns):
            sheet.write(0, column_index, column_name, header_style)
            width = min(max(len(column_name) + 4, 15), 55)
            sheet.col(column_index).width = width * 256

        for row_index, row in enumerate(dataframe.itertuples(index=False), start=1):
            for column_index, value in enumerate(row):
                if isinstance(value, bool):
                    sheet.write(row_index, column_index, str(value), text_style)
                elif isinstance(value, int):
                    sheet.write(row_index, column_index, value, number_style)
                elif isinstance(value, float):
                    sheet.write(row_index, column_index, value, money_style)
                else:
                    sheet.write(
                        row_index,
                        column_index,
                        "" if value is None else str(value),
                        text_style,
                    )
        sheet.set_panes_frozen(True)
        sheet.set_horz_split_pos(1)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def build_quote_xlsx(
    rate,
    profit_rate_percent,
    prepared_by,
    quote_date,
    quote_number,
    email_address,
):
    frames = build_export_frames(rate)
    dataframe = frames["Liste"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teklif"

    accent_fill = PatternFill(
        "solid",
        fgColor="C00000",
    )
    dark_fill = PatternFill(
        "solid",
        fgColor="1F2937",
    )
    light_fill = PatternFill(
        "solid",
        fgColor="E5E7EB",
    )
    total_fill = PatternFill(
        "solid",
        fgColor="FEE2E2",
    )
    white_font = Font(
        color="FFFFFF",
        bold=True,
    )
    title_font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )
    label_font = Font(
        bold=True,
        color="1F2937",
    )
    total_font = Font(
        bold=True,
        color="991B1B",
    )
    thin_gray = Side(
        style="thin",
        color="D1D5DB",
    )
    cell_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    sheet.merge_cells("A1:O1")
    title_cell = sheet["A1"]
    title_cell.value = "ITS SYSTEMS – TEKLİF"
    title_cell.fill = accent_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    sheet.row_dimensions[1].height = 30

    # Teklif üst bilgileri
    metadata = [
        ("A3", "Teklif Hazırlayan", "B3:D3", prepared_by),
        ("E3", "E-posta", "F3:I3", email_address),
        ("J3", "Teklif Tarihi", "K3:L3", quote_date),
        ("M3", "Teklif Numarası", "N3:O3", quote_number),
        ("A4", "Kâr Oranı", "B4:D4", float(profit_rate_percent) / 100),
        ("E4", "EUR/TL Kuru", "F4:I4", float(rate)),
    ]

    for label_cell, label, value_range, value in metadata:
        sheet[label_cell] = label
        sheet[label_cell].font = label_font
        sheet[label_cell].fill = light_fill
        sheet[label_cell].border = cell_border
        sheet[label_cell].alignment = Alignment(
            vertical="center",
        )

        sheet.merge_cells(value_range)
        start_cell = value_range.split(":")[0]
        sheet[start_cell] = value
        sheet[start_cell].border = cell_border
        sheet[start_cell].alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

        start_col = sheet[start_cell].column
        end_col = sheet[value_range.split(":")[1]].column
        row_number = sheet[start_cell].row

        for column_number in range(start_col, end_col + 1):
            sheet.cell(
                row=row_number,
                column=column_number,
            ).border = cell_border

    sheet["B4"].number_format = "0.00%"
    sheet["F4"].number_format = '#,##0.0000'
    sheet["K3"].number_format = "dd.mm.yyyy"

    headers = [
        "Sıra",
        "Parça Adı",
        "Adet",
        "Ebat (mm)",
        "Ağırlık (kg)",
        "İşlemler",
        "Birim Maliyet EUR",
        "Birim Maliyet TL",
        "Toplam Maliyet EUR",
        "Toplam Maliyet TL",
        "Kâr Oranı",
        "Kâr EUR",
        "Kâr TL",
        "Teklif Toplam EUR",
        "Teklif Toplam TL",
    ]
    header_row = 7

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )
        cell.fill = dark_fill
        cell.font = white_font
        cell.border = cell_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    data_start_row = header_row + 1

    if dataframe.empty:
        sheet.merge_cells(
            start_row=data_start_row,
            start_column=1,
            end_row=data_start_row,
            end_column=15,
        )
        empty_cell = sheet.cell(
            row=data_start_row,
            column=1,
            value="Teklife eklenecek kayıtlı parça bulunmuyor.",
        )
        empty_cell.alignment = Alignment(
            horizontal="center",
        )
        empty_cell.border = cell_border
        total_row = data_start_row + 1
    else:
        for row_offset, (_, row) in enumerate(
            dataframe.iterrows(),
            start=0,
        ):
            excel_row = data_start_row + row_offset

            values = [
                row_offset + 1,
                str(row["Parça Adı"]),
                int(row["Adet"]),
                str(row["Kaba Ebat"]),
                float(row["Malzeme Ağırlığı (kg)"]),
                str(row["İşlemler"]),
                float(row["Birim Fiyat EUR"]),
                float(row["Birim Fiyat TL"]),
                float(row["Toplam Fiyat EUR"]),
                float(row["Toplam Fiyat TL"]),
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                cell = sheet.cell(
                    row=excel_row,
                    column=column_index,
                    value=value,
                )
                cell.border = cell_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            sheet.cell(
                row=excel_row,
                column=11,
                value="=$B$4",
            )
            sheet.cell(
                row=excel_row,
                column=12,
                value=f"=I{excel_row}*K{excel_row}",
            )
            sheet.cell(
                row=excel_row,
                column=13,
                value=f"=J{excel_row}*K{excel_row}",
            )
            sheet.cell(
                row=excel_row,
                column=14,
                value=f"=I{excel_row}+L{excel_row}",
            )
            sheet.cell(
                row=excel_row,
                column=15,
                value=f"=J{excel_row}+M{excel_row}",
            )

            for column_index in range(11, 16):
                cell = sheet.cell(
                    row=excel_row,
                    column=column_index,
                )
                cell.border = cell_border
                cell.alignment = Alignment(
                    vertical="top",
                )

        total_row = data_start_row + len(dataframe)

    # Genel toplam satırı
    sheet.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=8,
    )
    total_label = sheet.cell(
        row=total_row,
        column=1,
        value="GENEL TOPLAM",
    )
    total_label.fill = total_fill
    total_label.font = total_font
    total_label.alignment = Alignment(
        horizontal="right",
        vertical="center",
    )

    for column_index in range(1, 16):
        sheet.cell(
            row=total_row,
            column=column_index,
        ).border = cell_border
        sheet.cell(
            row=total_row,
            column=column_index,
        ).fill = total_fill
        sheet.cell(
            row=total_row,
            column=column_index,
        ).font = total_font

    if not dataframe.empty:
        last_data_row = total_row - 1
        sheet.cell(
            row=total_row,
            column=9,
            value=f"=SUM(I{data_start_row}:I{last_data_row})",
        )
        sheet.cell(
            row=total_row,
            column=10,
            value=f"=SUM(J{data_start_row}:J{last_data_row})",
        )
        sheet.cell(
            row=total_row,
            column=11,
            value="=$B$4",
        )
        sheet.cell(
            row=total_row,
            column=12,
            value=f"=SUM(L{data_start_row}:L{last_data_row})",
        )
        sheet.cell(
            row=total_row,
            column=13,
            value=f"=SUM(M{data_start_row}:M{last_data_row})",
        )
        sheet.cell(
            row=total_row,
            column=14,
            value=f"=SUM(N{data_start_row}:N{last_data_row})",
        )
        sheet.cell(
            row=total_row,
            column=15,
            value=f"=SUM(O{data_start_row}:O{last_data_row})",
        )

    # Sayısal biçimler
    eur_columns = (7, 9, 12, 14)
    tl_columns = (8, 10, 13, 15)

    for row_number in range(data_start_row, total_row + 1):
        sheet.cell(
            row=row_number,
            column=3,
        ).number_format = "0"
        sheet.cell(
            row=row_number,
            column=5,
        ).number_format = "#,##0.0000"

        for column_index in eur_columns:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = '#,##0.00 "€"'

        for column_index in tl_columns:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = '#,##0.00 "TL"'

        sheet.cell(
            row=row_number,
            column=11,
        ).number_format = "0.00%"

    widths = {
        "A": 7,
        "B": 24,
        "C": 9,
        "D": 20,
        "E": 14,
        "F": 55,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 12,
        "L": 15,
        "M": 15,
        "N": 19,
        "O": 19,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.row_dimensions[header_row].height = 36
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = (
        f"A{header_row}:O{max(total_row - 1, header_row)}"
    )
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:O{total_row}"

    # Excel açıldığında formüllerin yeniden hesaplanmasını ister.
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================
# UYGULAMA VERİLERİ
# =========================================================
settings = get_settings()
exchange_rate = float(settings.get("eur_tl_kuru") or 50)

prices = get_prices()
labors = get_labors()
parts = get_parts()

price_map = {item["id"]: item for item in prices}
labor_map = {item["id"]: item for item in labors}

render_header(settings)
render_rate(settings)

(
    tab_suppliers,
    tab_prices,
    tab_labors,
    tab_cost,
    tab_list,
) = st.tabs(
    [
        "Tedarikçi Listesi",
        "Fiyat Tanımları",
        "İşçilik Maliyetleri",
        "Parça Maliyeti",
        "Teklif",
    ]
)


# =========================================================
# FİYAT TANIMLARI
# =========================================================
with tab_prices:
    st.subheader("Yeni fiyat tanımı")

    with st.form("new_price_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            new_category = st.selectbox(
                "Kategori", ["Malzeme", "Kaplama", "Ek İşlem"]
            )
            new_name = st.text_input("Ad")
            new_description = st.text_input("Açıklama")
        with col2:
            new_currency = st.selectbox("Fiyat para birimi", ["EUR", "TL"])
            new_price_text = st.text_input(
                f"Birim fiyat ({new_currency}) — Malzemede kg fiyatı",
                value="0,00",
            )
            new_density_text = st.text_input(
                "Yoğunluk (g/cm³) — yalnızca malzeme için",
                value="",
            )

        save_new_price = st.form_submit_button(
            "Fiyat tanımını kaydet", type="primary"
        )
        if save_new_price:
            new_price = parse_decimal(new_price_text)
            new_density = parse_decimal(new_density_text, 0)
            if not new_name.strip():
                st.error("Ad alanı boş bırakılamaz.")
            elif new_price is None or new_price < 0:
                st.error("Geçerli bir fiyat gir.")
            elif new_category == "Malzeme" and (new_density is None or new_density <= 0):
                st.error("Malzeme için yoğunluk girmelisin.")
            else:
                eur_snapshot, _ = convert_price(
                    new_price, new_currency, exchange_rate
                )
                db.table("fiyat_tanimlari").insert({
                    "kategori": new_category,
                    "ad": new_name.strip(),
                    "aciklama": new_description.strip(),
                    "kaynak_para_birimi": new_currency,
                    "kaynak_birim_fiyat": new_price,
                    "birim_fiyat_eur": eur_snapshot,
                    "yogunluk_g_cm3": (
                        new_density if new_category == "Malzeme" else None
                    ),
                }).execute()
                st.rerun()

    st.divider()
    st.subheader("Düzenle veya sil")
    selected_filter = st.selectbox(
        "Kategoriye göre filtrele",
        ["Tümü", "Malzeme", "Kaplama", "Ek İşlem"],
        key="price_filter",
    )
    filtered_prices = prices if selected_filter == "Tümü" else [
        item for item in prices if item["kategori"] == selected_filter
    ]

    if not filtered_prices:
        st.info("Bu filtreye uygun kayıt bulunmuyor.")
    else:
        for item in filtered_prices:
            currency, source_value = get_price_source(item)
            eur_value, tl_value = convert_price(
                source_value, currency, exchange_rate
            )
            density = get_density(item)
            title = (
                f'{item["ad"]} | {item["kategori"]} | '
                f'{format_eur(eur_value)} / {format_tl(tl_value)}'
            )
            if item["kategori"] == "Malzeme":
                title += f' | {format_number(density, 4)} g/cm³'

            with st.expander(title, expanded=False):
                with st.form(f'price_edit_form_{item["id"]}'):
                    col1, col2 = st.columns(2)
                    with col1:
                        categories = ["Malzeme", "Kaplama", "Ek İşlem"]
                        edit_category = st.selectbox(
                            "Kategori",
                            categories,
                            index=categories.index(item["kategori"]),
                        )
                        edit_name = st.text_input("Ad", value=item["ad"])
                        edit_description = st.text_input(
                            "Açıklama", value=item.get("aciklama", "") or ""
                        )
                    with col2:
                        edit_currency = st.selectbox(
                            "Fiyat para birimi",
                            ["EUR", "TL"],
                            index=["EUR", "TL"].index(currency),
                        )
                        edit_price_text = st.text_input(
                            f"Birim fiyat ({edit_currency}) — Malzemede kg fiyatı",
                            value=format_number(source_value, 4),
                        )
                        edit_density_text = st.text_input(
                            "Yoğunluk (g/cm³) — yalnızca malzeme için",
                            value=format_number(density, 4) if density > 0 else "",
                        )

                    update_col, delete_col = st.columns(2)
                    with update_col:
                        update_price = st.form_submit_button(
                            "Güncelle", type="primary", use_container_width=True
                        )
                    with delete_col:
                        delete_price = st.form_submit_button(
                            "Sil", use_container_width=True
                        )

                    if update_price:
                        edit_price = parse_decimal(edit_price_text)
                        edit_density = parse_decimal(edit_density_text, 0)
                        if not edit_name.strip():
                            st.error("Ad alanı boş bırakılamaz.")
                        elif edit_price is None or edit_price < 0:
                            st.error("Geçerli bir fiyat gir.")
                        elif edit_category == "Malzeme" and (
                            edit_density is None or edit_density <= 0
                        ):
                            st.error("Malzeme için yoğunluk girmelisin.")
                        else:
                            eur_snapshot, _ = convert_price(
                                edit_price, edit_currency, exchange_rate
                            )
                            db.table("fiyat_tanimlari").update({
                                "kategori": edit_category,
                                "ad": edit_name.strip(),
                                "aciklama": edit_description.strip(),
                                "kaynak_para_birimi": edit_currency,
                                "kaynak_birim_fiyat": edit_price,
                                "birim_fiyat_eur": eur_snapshot,
                                "yogunluk_g_cm3": (
                                    edit_density if edit_category == "Malzeme" else None
                                ),
                            }).eq("id", item["id"]).execute()
                            st.rerun()

                    if delete_price:
                        try:
                            db.table("fiyat_tanimlari").delete().eq(
                                "id", item["id"]
                            ).execute()
                            st.rerun()
                        except Exception:
                            st.error("Bu kayıt bir parçada kullanıldığı için silinemedi.")
# =========================================================
# İŞÇİLİK MALİYETLERİ
# =========================================================
with tab_labors:
    st.subheader("Yeni işçilik maliyeti")

    with st.form("new_labor_form", clear_on_submit=True):
        col1, col2 = st.columns(2)

        with col1:
            labor_name = st.text_input(
                "İşçilik adı",
                placeholder="Örn. Dik İşlem CNC, Torna",
            )
            labor_description = st.text_input("Açıklama")

        with col2:
            labor_currency = st.selectbox(
                "Saatlik ücret para birimi",
                ["EUR", "TL"],
            )
            labor_price_text = st.text_input(
                f"Saatlik ücret ({labor_currency})",
                value="0,00",
            )

        save_labor = st.form_submit_button(
            "İşçilik tanımını kaydet",
            type="primary",
        )

        if save_labor:
            labor_price = parse_decimal(labor_price_text)

            if not labor_name.strip():
                st.error("İşçilik adı boş bırakılamaz.")
            elif labor_price is None or labor_price < 0:
                st.error("Geçerli bir saatlik ücret gir.")
            else:
                db.table("iscilik_tanimlari").insert(
                    {
                        "ad": labor_name.strip(),
                        "aciklama": labor_description.strip(),
                        "kaynak_para_birimi": labor_currency,
                        "kaynak_saatlik_ucret": labor_price,
                    }
                ).execute()

                st.success("İşçilik tanımı kaydedildi.")
                st.rerun()

    st.divider()
    st.subheader("Düzenle veya sil")

    if not labors:
        st.info("Henüz işçilik tanımı bulunmuyor.")
    else:
        for item in labors:
            currency, source_value = get_labor_source(item)
            eur_value, tl_value = convert_price(
                source_value,
                currency,
                exchange_rate,
            )

            title = (
                f'{item["ad"]} | '
                f'{format_eur(eur_value)}/saat | '
                f'{format_tl(tl_value)}/saat'
            )

            with st.expander(title, expanded=False):
                with st.form(f'labor_edit_form_{item["id"]}'):
                    col1, col2 = st.columns(2)

                    with col1:
                        edit_labor_name = st.text_input(
                            "İşçilik adı",
                            value=item["ad"],
                        )
                        edit_labor_description = st.text_input(
                            "Açıklama",
                            value=item.get("aciklama", "") or "",
                        )

                    with col2:
                        edit_labor_currency = st.selectbox(
                            "Saatlik ücret para birimi",
                            ["EUR", "TL"],
                            index=["EUR", "TL"].index(currency),
                        )
                        edit_labor_price_text = st.text_input(
                            f"Saatlik ücret ({edit_labor_currency})",
                            value=format_number(source_value, 4),
                        )

                    update_col, delete_col = st.columns(2)

                    with update_col:
                        update_labor = st.form_submit_button(
                            "Güncelle",
                            type="primary",
                            use_container_width=True,
                        )

                    with delete_col:
                        delete_labor = st.form_submit_button(
                            "Sil",
                            use_container_width=True,
                        )

                    if update_labor:
                        edit_labor_price = parse_decimal(
                            edit_labor_price_text
                        )

                        if not edit_labor_name.strip():
                            st.error("İşçilik adı boş bırakılamaz.")
                        elif (
                            edit_labor_price is None
                            or edit_labor_price < 0
                        ):
                            st.error("Geçerli bir saatlik ücret gir.")
                        else:
                            db.table("iscilik_tanimlari").update(
                                {
                                    "ad": edit_labor_name.strip(),
                                    "aciklama": edit_labor_description.strip(),
                                    "kaynak_para_birimi": edit_labor_currency,
                                    "kaynak_saatlik_ucret": edit_labor_price,
                                }
                            ).eq("id", item["id"]).execute()

                            st.success("İşçilik güncellendi.")
                            st.rerun()

                    if delete_labor:
                        try:
                            db.table("iscilik_tanimlari").delete().eq(
                                "id",
                                item["id"],
                            ).execute()
                            st.success("İşçilik silindi.")
                            st.rerun()
                        except Exception:
                            st.error(
                                "Bu işçilik bir parçada kullanıldığı için silinemedi."
                            )


# =========================================================
# LİSTE
# =========================================================
# =========================================================
# PARÇA MALİYETİ
# =========================================================
# =========================================================
# PARÇA MALİYETİ
# =========================================================
with tab_cost:
    st.subheader("Parça Maliyeti")
    st.caption(
        "Tüm bilgileri gir. Hesaplama yalnızca Güncelle "
        "butonuna bastığında yapılır."
    )

    if st.session_state.pop("part_saved_success", False):
        st.success(
            "Parça detaylarıyla birlikte Liste bölümüne kaydedildi. "
            "Form yeni parça girişi için sıfırlandı."
        )

    form_version = int(
        st.session_state.get("part_form_version", 1)
    )
    context_id = f"new_{form_version}"
    preview_key = f"part_preview_{context_id}"

    materials = [
        item
        for item in prices
        if item["kategori"] == "Malzeme"
    ]
    coatings = [
        item
        for item in prices
        if item["kategori"] == "Kaplama"
    ]
    extras = [
        item
        for item in prices
        if item["kategori"] == "Ek İşlem"
    ]

    measurement_labors = [
        item
        for item in labors
        if "ölçüm" in item["ad"].casefold()
    ]
    machining_labors = [
        item
        for item in labors
        if "ölçüm" not in item["ad"].casefold()
    ]

    material_map = {
        item["id"]: item
        for item in materials
    }
    coating_map = {
        item["id"]: item
        for item in coatings
    }
    extra_map = {
        item["id"]: item
        for item in extras
    }
    measurement_map = {
        item["id"]: item
        for item in measurement_labors
    }

    material_ids = [None] + list(material_map.keys())
    coating_ids = [None] + list(coating_map.keys())
    extra_ids = [None] + list(extra_map.keys())
    measurement_ids = [None] + list(
        measurement_map.keys()
    )

    def material_label(item_id):
        if item_id is None:
            return "Seçiniz"

        item = material_map[item_id]
        currency, source_value = get_price_source(item)
        eur_value, tl_value = convert_price(
            source_value,
            currency,
            exchange_rate,
        )
        density = get_density(item)

        return (
            f'{item["ad"]} — '
            f'{format_eur(eur_value)}/kg / '
            f'{format_tl(tl_value)}/kg — '
            f'{format_number(density, 4)} g/cm³'
        )

    def price_label(item_id, item_map):
        if item_id is None:
            return "Seçiniz"

        item = item_map[item_id]
        currency, source_value = get_price_source(item)
        eur_value, tl_value = convert_price(
            source_value,
            currency,
            exchange_rate,
        )

        return (
            f'{item["ad"]} — '
            f'{format_eur(eur_value)} / '
            f'{format_tl(tl_value)}'
        )

    def labor_label(item_id, item_map):
        if item_id is None:
            return "Seçiniz"

        item = item_map[item_id]
        currency, source_value = get_labor_source(item)
        eur_value, tl_value = convert_price(
            source_value,
            currency,
            exchange_rate,
        )

        return (
            f'{item["ad"]} — '
            f'{format_eur(eur_value)}/saat / '
            f'{format_tl(tl_value)}/saat'
        )

    machining_label_map = {}
    machining_options = []

    for item in machining_labors:
        label = labor_label(
            item["id"],
            {item["id"]: item},
        )
        if label in machining_label_map:
            label = f'{label} (ID: {item["id"]})'
        machining_options.append(label)
        machining_label_map[label] = item

    machining_seed = pd.DataFrame(
        [
            {
                "Talaşlı İmalat": None,
                "Süre": None,
                "Birim": "Saat",
            }
        ]
    )

    preview = st.session_state.get(preview_key)

    with st.form(
        key=f"part_cost_form_{context_id}",
        clear_on_submit=False,
    ):
        top_col1, top_col2 = st.columns([3, 1])

        with top_col1:
            part_name = st.text_input(
                "Parça adı",
                key=f"part_name_{context_id}",
            )

        with top_col2:
            production_quantity = st.number_input(
                "Üretilecek adet",
                min_value=1,
                value=1,
                step=1,
                key=f"production_quantity_{context_id}",
            )

        st.markdown("### 1. Malzeme ve ölçüler")

        (
            material_col,
            length_col,
            width_col,
            height_col,
        ) = st.columns([2.8, 1, 1, 1.15])

        with material_col:
            selected_material_id = st.selectbox(
                "Malzeme",
                material_ids,
                format_func=material_label,
                disabled=len(material_ids) == 1,
                key=f"material_select_{context_id}",
            )

        with length_col:
            length_mm = st.number_input(
                "Boy (mm)",
                min_value=0.0,
                value=0.0,
                step=1.0,
                format="%.3f",
                key=f"length_mm_{context_id}",
            )

        with width_col:
            width_mm = st.number_input(
                "En (mm)",
                min_value=0.0,
                value=0.0,
                step=1.0,
                format="%.3f",
                key=f"width_mm_{context_id}",
            )

        with height_col:
            height_mm = st.number_input(
                "Yükseklik / Kalınlık (mm)",
                min_value=0.0,
                value=0.0,
                step=1.0,
                format="%.3f",
                key=f"height_mm_{context_id}",
            )

        st.caption(
            "Malzeme bedeli ve ağırlık Güncelle butonundan "
            "sonra hesaplanır."
        )

        st.divider()
        st.markdown("### 2. Talaşlı imalat")
        st.caption(
            "Tablonun altındaki + işaretiyle birden fazla satır "
            "ekleyebilirsin. Dakika seçilirse süre otomatik "
            "olarak saate çevrilir."
        )

        machining_editor = st.data_editor(
            machining_seed,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key=f"machining_editor_{context_id}",
            column_config={
                "Talaşlı İmalat": st.column_config.SelectboxColumn(
                    "Talaşlı İmalat",
                    options=machining_options,
                    required=False,
                    width="large",
                ),
                "Süre": st.column_config.NumberColumn(
                    "Süre",
                    min_value=0.0,
                    step=1.0,
                    format="%.4f",
                ),
                "Birim": st.column_config.SelectboxColumn(
                    "Birim",
                    options=["Saat", "Dakika"],
                    required=True,
                    width="small",
                ),
            },
            disabled=(
                len(machining_options) == 0
            ),
        )

        if not machining_options:
            st.info(
                "İşçilik Maliyetleri bölümünde talaşlı imalat "
                "kaydı bulunmuyor."
            )

        st.divider()
        st.markdown("### 3. Kaplama")

        coating_col1, coating_col2 = st.columns([4, 1])

        with coating_col1:
            selected_coating_id = st.selectbox(
                "Kaplama",
                coating_ids,
                format_func=lambda item_id: price_label(
                    item_id,
                    coating_map,
                ),
                disabled=len(coating_ids) == 1,
                key=f"coating_select_{context_id}",
            )

        with coating_col2:
            coating_quantity = st.number_input(
                "Kaplama adedi",
                min_value=1,
                value=1,
                step=1,
                key=f"coating_quantity_{context_id}",
            )

        st.divider()
        st.markdown("### 4. Ek işlem")

        (
            extra_col1,
            extra_col2,
            extra_col3,
        ) = st.columns([3.5, 1, 1])

        with extra_col1:
            selected_extra_id = st.selectbox(
                "Ek İşlem",
                extra_ids,
                format_func=lambda item_id: price_label(
                    item_id,
                    extra_map,
                ),
                disabled=len(extra_ids) == 1,
                key=f"extra_select_{context_id}",
            )

        with extra_col2:
            extra_unit = st.selectbox(
                "Birim",
                ["Adet", "Saat"],
                key=f"extra_unit_{context_id}",
            )

        with extra_col3:
            extra_amount = st.number_input(
                "Miktar",
                min_value=0.0,
                value=1.0,
                step=0.25,
                format="%.4f",
                key=f"extra_amount_{context_id}",
            )

        if len(extra_ids) == 1:
            st.caption(
                "Ek İşlem seçeneği bulunmuyor. Fiyat Tanımları "
                "bölümünden Ek İşlem ekleyebilirsin."
            )

        st.divider()
        st.markdown("### 5. Ölçüm")

        measurement_col1, measurement_col2 = st.columns([4, 1])

        with measurement_col1:
            selected_measurement_id = st.selectbox(
                "Ölçüm",
                measurement_ids,
                format_func=lambda item_id: labor_label(
                    item_id,
                    measurement_map,
                ),
                disabled=len(measurement_ids) == 1,
                key=f"measurement_select_{context_id}",
            )

        with measurement_col2:
            measurement_hours = st.number_input(
                "Süre (saat)",
                min_value=0.0,
                value=1.0,
                step=0.25,
                format="%.4f",
                key=f"measurement_hours_{context_id}",
            )

        if len(measurement_ids) == 1:
            st.caption(
                'Ölçüm seçeneği bulunmuyor. İşçilik Maliyetleri '
                'bölümüne adı "Ölçüm" içeren bir kayıt '
                'ekleyebilirsin.'
            )

        st.divider()
        update_col, save_col = st.columns(2)

        with update_col:
            update_clicked = st.form_submit_button(
                "Güncelle",
                use_container_width=True,
            )

        with save_col:
            save_clicked = st.form_submit_button(
                "Parçayı Kaydet",
                use_container_width=True,
                disabled=preview is None,
            )

    selected_material = (
        material_map.get(selected_material_id)
        if selected_material_id is not None
        else None
    )
    selected_coating = (
        coating_map.get(selected_coating_id)
        if selected_coating_id is not None
        else None
    )
    selected_extra = (
        extra_map.get(selected_extra_id)
        if selected_extra_id is not None
        else None
    )
    selected_measurement = (
        measurement_map.get(selected_measurement_id)
        if selected_measurement_id is not None
        else None
    )

    machining_rows = []
    machining_error = None

    for row_index, row in machining_editor.iterrows():
        operation_value = row.get("Talaşlı İmalat")
        duration_value = row.get("Süre")
        unit_value = row.get("Birim")

        operation_label = (
            ""
            if pd.isna(operation_value)
            else str(operation_value).strip()
        )
        duration = (
            0.0
            if pd.isna(duration_value)
            else float(duration_value)
        )
        duration_unit = (
            "Saat"
            if pd.isna(unit_value)
            else str(unit_value)
        )

        if not operation_label and duration <= 0:
            continue

        if not operation_label and duration > 0:
            machining_error = (
                f"Talaşlı imalat tablosundaki {row_index + 1}. "
                "satırda işlem seçmeden süre girdin."
            )
            break

        selected_labor = machining_label_map.get(
            operation_label
        )

        if selected_labor is None:
            machining_error = (
                f"Talaşlı imalat tablosundaki {row_index + 1}. "
                "satırın işlemi geçersiz."
            )
            break

        if duration <= 0:
            machining_error = (
                f"Talaşlı imalat tablosundaki {row_index + 1}. "
                "satırın süresini sıfırdan büyük gir."
            )
            break

        labor_hours = (
            duration / 60.0
            if duration_unit == "Dakika"
            else duration
        )
        labor_currency, labor_source = get_labor_source(
            selected_labor
        )

        machining_rows.append(
            {
                "definition": selected_labor,
                "hours": labor_hours,
                "entered_value": duration,
                "entered_unit": duration_unit,
                "currency": labor_currency,
                "source_value": labor_source,
            }
        )

    material_row = None

    if selected_material is not None:
        material_currency, material_source = get_price_source(
            selected_material
        )
        material_row = {
            "definition": selected_material,
            "currency": material_currency,
            "source_value": material_source,
            "density": get_density(selected_material),
            "length_mm": float(length_mm),
            "width_mm": float(width_mm),
            "height_mm": float(height_mm),
        }

    coating_row = None

    if selected_coating is not None:
        coating_currency, coating_source = get_price_source(
            selected_coating
        )
        coating_row = {
            "definition": selected_coating,
            "quantity": int(coating_quantity),
            "amount_type": "adet",
            "currency": coating_currency,
            "source_value": coating_source,
        }

    extra_row = None

    if selected_extra is not None:
        extra_currency, extra_source = get_price_source(
            selected_extra
        )
        extra_row = {
            "definition": selected_extra,
            "quantity": float(extra_amount),
            "amount_type": (
                "adet"
                if extra_unit == "Adet"
                else "saat"
            ),
            "currency": extra_currency,
            "source_value": extra_source,
        }

    measurement_row = None

    if selected_measurement is not None:
        (
            measurement_currency,
            measurement_source,
        ) = get_labor_source(selected_measurement)
        measurement_row = {
            "definition": selected_measurement,
            "hours": float(measurement_hours),
            "entered_value": float(measurement_hours),
            "entered_unit": "Saat",
            "currency": measurement_currency,
            "source_value": measurement_source,
        }

    def create_draft():
        single_eur = 0.0
        single_tl = 0.0
        calculated_material = None
        calculated_operations = []
        calculated_measurement = None
        calculated_labors = []

        if material_row is not None:
            (
                volume_mm3,
                volume_cm3,
                weight_kg,
            ) = calculate_rectangular_material(
                material_row["length_mm"],
                material_row["width_mm"],
                material_row["height_mm"],
                material_row["density"],
            )

            unit_eur, unit_tl = convert_price(
                material_row["source_value"],
                material_row["currency"],
                exchange_rate,
            )
            line_eur = unit_eur * weight_kg
            line_tl = unit_tl * weight_kg
            single_eur += line_eur
            single_tl += line_tl

            calculated_material = {
                **material_row,
                "weight_kg": weight_kg,
                "volume_mm3": volume_mm3,
                "volume_cm3": volume_cm3,
                "line_eur": line_eur,
                "line_tl": line_tl,
            }

        for operation_row in (
            coating_row,
            extra_row,
        ):
            if operation_row is None:
                continue

            unit_eur, unit_tl = convert_price(
                operation_row["source_value"],
                operation_row["currency"],
                exchange_rate,
            )
            line_eur = (
                unit_eur * operation_row["quantity"]
            )
            line_tl = (
                unit_tl * operation_row["quantity"]
            )
            single_eur += line_eur
            single_tl += line_tl

            calculated_operations.append(
                {
                    **operation_row,
                    "line_eur": line_eur,
                    "line_tl": line_tl,
                }
            )

        if measurement_row is not None:
            hourly_eur, hourly_tl = convert_price(
                measurement_row["source_value"],
                measurement_row["currency"],
                exchange_rate,
            )
            line_eur = (
                hourly_eur * measurement_row["hours"]
            )
            line_tl = (
                hourly_tl * measurement_row["hours"]
            )
            single_eur += line_eur
            single_tl += line_tl

            calculated_measurement = {
                **measurement_row,
                "line_eur": line_eur,
                "line_tl": line_tl,
            }

        for labor_row in machining_rows:
            hourly_eur, hourly_tl = convert_price(
                labor_row["source_value"],
                labor_row["currency"],
                exchange_rate,
            )
            line_eur = hourly_eur * labor_row["hours"]
            line_tl = hourly_tl * labor_row["hours"]
            single_eur += line_eur
            single_tl += line_tl

            calculated_labors.append(
                {
                    **labor_row,
                    "line_eur": line_eur,
                    "line_tl": line_tl,
                }
            )

        signature = (
            part_name.strip(),
            int(production_quantity),
            round(float(exchange_rate), 8),
            (
                None
                if material_row is None
                else (
                    int(material_row["definition"]["id"]),
                    round(material_row["length_mm"], 6),
                    round(material_row["width_mm"], 6),
                    round(material_row["height_mm"], 6),
                    round(material_row["density"], 6),
                )
            ),
            tuple(
                (
                    int(
                        operation_row[
                            "definition"
                        ]["id"]
                    ),
                    round(
                        float(operation_row["quantity"]),
                        8,
                    ),
                    operation_row["amount_type"],
                )
                for operation_row in (
                    coating_row,
                    extra_row,
                )
                if operation_row is not None
            ),
            (
                None
                if measurement_row is None
                else (
                    int(
                        measurement_row[
                            "definition"
                        ]["id"]
                    ),
                    round(measurement_row["hours"], 8),
                )
            ),
            tuple(
                (
                    int(labor_row["definition"]["id"]),
                    round(labor_row["hours"], 8),
                    labor_row["entered_unit"],
                    round(labor_row["entered_value"], 8),
                )
                for labor_row in machining_rows
            ),
        )

        return {
            "part_name": part_name.strip(),
            "production_quantity": int(
                production_quantity
            ),
            "material": calculated_material,
            "operation_rows": calculated_operations,
            "measurement": calculated_measurement,
            "labors": calculated_labors,
            "single_eur": single_eur,
            "single_tl": single_tl,
            "total_eur": (
                single_eur * int(production_quantity)
            ),
            "total_tl": (
                single_tl * int(production_quantity)
            ),
            "signature": signature,
        }

    submitted = update_clicked or save_clicked
    current_draft = create_draft() if submitted else None

    validation_error = None

    if submitted:
        if not part_name.strip():
            validation_error = "Parça adı boş bırakılamaz."
        elif selected_material is None:
            validation_error = "Bir malzeme seçmelisin."
        elif material_row["density"] <= 0:
            validation_error = (
                "Seçilen malzemenin yoğunluğu tanımlanmamış. "
                "Fiyat Tanımları bölümünden yoğunluğu güncelle."
            )
        elif any(
            value <= 0
            for value in (
                material_row["length_mm"],
                material_row["width_mm"],
                material_row["height_mm"],
            )
        ):
            validation_error = (
                "Boy, en ve yükseklik değerlerinin tamamını "
                "sıfırdan büyük gir."
            )
        elif machining_error:
            validation_error = machining_error
        elif (
            selected_measurement is not None
            and float(measurement_hours) <= 0
        ):
            validation_error = (
                "Ölçüm süresini sıfırdan büyük gir."
            )
        elif (
            selected_extra is not None
            and float(extra_amount) <= 0
        ):
            validation_error = (
                "Ek işlem miktarını sıfırdan büyük gir."
            )

    if validation_error:
        st.error(validation_error)

    if update_clicked and not validation_error:
        st.session_state[preview_key] = current_draft
        st.rerun()

    if save_clicked and not validation_error:
        saved_preview = st.session_state.get(preview_key)

        if saved_preview is None:
            st.error(
                "Önce Güncelle butonuna basarak fiyatı hesapla."
            )
        elif (
            current_draft["signature"]
            != saved_preview["signature"]
        ):
            st.error(
                "Bilgiler hesaplamadan sonra değişmiş. "
                "Önce tekrar Güncelle butonuna bas."
            )
        else:
            material = saved_preview["material"]

            result = db.table("parcalar").insert(
                {
                    "parca_adi": saved_preview["part_name"],
                    "adet": saved_preview[
                        "production_quantity"
                    ],
                    "boy_mm": material["length_mm"],
                    "en_mm": material["width_mm"],
                    "yukseklik_mm": material["height_mm"],
                    "malzeme_agirlik_kg": material[
                        "weight_kg"
                    ],
                }
            ).execute()

            part_id = result.data[0]["id"]
            item_rows_to_save = []

            material_eur_snapshot, _ = convert_price(
                material["source_value"],
                material["currency"],
                exchange_rate,
            )

            item_rows_to_save.append(
                {
                    "parca_id": part_id,
                    "fiyat_tanimi_id": material[
                        "definition"
                    ]["id"],
                    "miktar": material["weight_kg"],
                    "miktar_turu": "kg",
                    "kaynak_para_birimi": material[
                        "currency"
                    ],
                    "kaynak_birim_fiyat": material[
                        "source_value"
                    ],
                    "birim_fiyat_eur": material_eur_snapshot,
                }
            )

            for operation_row in saved_preview[
                "operation_rows"
            ]:
                eur_snapshot, _ = convert_price(
                    operation_row["source_value"],
                    operation_row["currency"],
                    exchange_rate,
                )

                item_rows_to_save.append(
                    {
                        "parca_id": part_id,
                        "fiyat_tanimi_id": operation_row[
                            "definition"
                        ]["id"],
                        "miktar": operation_row["quantity"],
                        "miktar_turu": operation_row[
                            "amount_type"
                        ],
                        "kaynak_para_birimi": operation_row[
                            "currency"
                        ],
                        "kaynak_birim_fiyat": operation_row[
                            "source_value"
                        ],
                        "birim_fiyat_eur": eur_snapshot,
                    }
                )

            db.table("parca_kalemleri").insert(
                item_rows_to_save
            ).execute()

            labor_rows_to_save = []
            all_labor_rows = []

            if saved_preview["measurement"] is not None:
                all_labor_rows.append(
                    saved_preview["measurement"]
                )

            all_labor_rows.extend(saved_preview["labors"])

            for labor_item in all_labor_rows:
                labor_rows_to_save.append(
                    {
                        "parca_id": part_id,
                        "iscilik_tanimi_id": labor_item[
                            "definition"
                        ]["id"],
                        "saat": labor_item["hours"],
                        "kaynak_para_birimi": labor_item[
                            "currency"
                        ],
                        "kaynak_saatlik_ucret": labor_item[
                            "source_value"
                        ],
                    }
                )

            if labor_rows_to_save:
                db.table(
                    "parca_iscilik_kalemleri"
                ).insert(labor_rows_to_save).execute()

            st.session_state.pop(preview_key, None)
            st.session_state[
                "part_form_version"
            ] = form_version + 1
            st.session_state[
                "part_saved_success"
            ] = True
            st.rerun()

    preview = st.session_state.get(preview_key)

    if preview is not None:
        st.divider()
        st.markdown("### Hesaplanan Fiyat")

        material = preview["material"]

        material_info_col, single_col = st.columns(2)

        with material_info_col:
            st.metric(
                "Tek parça hammadde maliyeti",
                (
                    f'{format_eur(material["line_eur"])} '
                    f'/ {format_tl(material["line_tl"])}'
                ),
            )
            st.caption(
                f'(Hammadde ağırlığı: '
                f'{format_number(material["weight_kg"], 4)} kg, '
                f'hacim: '
                f'{format_number(material["volume_cm3"], 2)} cm³)'
            )

        with single_col:
            st.metric(
                "Tek parça toplam maliyet",
                (
                    f'{format_eur(preview["single_eur"])} '
                    f'/ {format_tl(preview["single_tl"])}'
                ),
            )

        total_col1, total_col2 = st.columns(2)
        total_col1.metric(
            "Genel toplam EUR",
            format_eur(preview["total_eur"]),
        )
        total_col2.metric(
            "Genel toplam TL",
            format_tl(preview["total_tl"]),
        )

        detail_rows = [
            {
                "Kalem": material["definition"]["ad"],
                "Miktar": (
                    f'{format_number(material["weight_kg"], 4)} kg'
                ),
                "Maliyet EUR": format_eur(
                    material["line_eur"]
                ),
                "Maliyet TL": format_tl(
                    material["line_tl"]
                ),
            }
        ]

        for operation_row in preview["operation_rows"]:
            amount_text = (
                f'{format_number(operation_row["quantity"], 4)} '
                f'saat'
                if operation_row["amount_type"] == "saat"
                else (
                    f'{format_number(operation_row["quantity"], 0)} '
                    f'adet'
                )
            )

            detail_rows.append(
                {
                    "Kalem": operation_row["definition"]["ad"],
                    "Miktar": amount_text,
                    "Maliyet EUR": format_eur(
                        operation_row["line_eur"]
                    ),
                    "Maliyet TL": format_tl(
                        operation_row["line_tl"]
                    ),
                }
            )

        labor_detail_rows = []

        if preview["measurement"] is not None:
            labor_detail_rows.append(preview["measurement"])

        labor_detail_rows.extend(preview["labors"])

        for labor_item in labor_detail_rows:
            detail_rows.append(
                {
                    "Kalem": labor_item["definition"]["ad"],
                    "Miktar": (
                        f'{format_number(labor_item["entered_value"], 4)} '
                        f'{labor_item["entered_unit"]} '
                        f'({format_number(labor_item["hours"], 4)} saat)'
                    ),
                    "Maliyet EUR": format_eur(
                        labor_item["line_eur"]
                    ),
                    "Maliyet TL": format_tl(
                        labor_item["line_tl"]
                    ),
                }
            )

        if detail_rows:
            st.dataframe(
                pd.DataFrame(detail_rows),
                use_container_width=True,
                hide_index=True,
            )

# =========================================================
# LİSTE
# =========================================================
with tab_list:
    st.subheader("Teklif")
    st.caption(
        "Kayıtlı parçaların maliyetlerini kontrol et, kâr oranını "
        "belirle ve teklif dosyasını Excel formatında oluştur."
    )

    deleted_count = st.session_state.pop(
        "deleted_all_parts_count",
        None,
    )
    if deleted_count:
        st.success(
            f"{deleted_count} parça ve bağlı maliyet kayıtları silindi."
        )

    all_part_items = get_part_items()
    all_part_labors = get_part_labors()

    if not parts:
        st.info("Henüz kayıtlı parça bulunmuyor.")
    else:
        list_rows = []
        list_total_cost_eur = 0.0
        list_total_cost_tl = 0.0

        for part in parts:
            part_id = part["id"]
            quantity = int(part["adet"])
            single_eur = 0.0
            single_tl = 0.0
            operations = []

            for row in [x for x in all_part_items if x["parca_id"] == part_id]:
                definition = price_map.get(row["fiyat_tanimi_id"], {})
                currency = row.get("kaynak_para_birimi") or "EUR"
                source_value = row.get("kaynak_birim_fiyat")
                if source_value is None:
                    source_value = row.get("birim_fiyat_eur", 0)
                unit_eur, unit_tl = convert_price(
                    float(source_value or 0), currency, exchange_rate
                )
                amount = float(row.get("miktar") or 0)
                amount_type = row.get("miktar_turu") or "adet"
                line_eur = unit_eur * amount
                line_tl = unit_tl * amount
                single_eur += line_eur
                single_tl += line_tl
                if amount_type == "kg":
                    amount_text = f"{format_number(amount, 4)} kg"
                elif amount_type == "saat":
                    amount_text = f"{format_number(amount, 4)} saat"
                else:
                    amount_text = f"{format_number(amount, 0)} adet"
                operations.append(
                    f'{definition.get("ad", "")} [{amount_text}] '
                    f'({format_eur(line_eur)} / {format_tl(line_tl)})'
                )

            for row in [x for x in all_part_labors if x["parca_id"] == part_id]:
                definition = labor_map.get(row["iscilik_tanimi_id"], {})
                currency = row.get("kaynak_para_birimi") or "EUR"
                source_value = float(row.get("kaynak_saatlik_ucret") or 0)
                hourly_eur, hourly_tl = convert_price(
                    source_value, currency, exchange_rate
                )
                hours = float(row.get("saat") or 0)
                line_eur = hourly_eur * hours
                line_tl = hourly_tl * hours
                single_eur += line_eur
                single_tl += line_tl
                operations.append(
                    f'{definition.get("ad", "")} [{format_number(hours, 4)} saat] '
                    f'({format_eur(line_eur)} / {format_tl(line_tl)})'
                )

            dimensions = ""
            if all(part.get(key) is not None for key in ("boy_mm", "en_mm", "yukseklik_mm")):
                dimensions = (
                    f'{format_number(part["boy_mm"], 2)} × '
                    f'{format_number(part["en_mm"], 2)} × '
                    f'{format_number(part["yukseklik_mm"], 2)}'
                )
            part_total_eur = single_eur * quantity
            part_total_tl = single_tl * quantity
            list_total_cost_eur += part_total_eur
            list_total_cost_tl += part_total_tl

            list_rows.append({
                "Parça Adı": part["parca_adi"],
                "Adet": quantity,
                "Ebat (mm)": dimensions,
                "Ağırlık (kg)": format_number(part.get("malzeme_agirlik_kg") or 0, 4),
                "İşlemler": " + ".join(operations),
                "Birim Fiyat EUR": format_eur(single_eur),
                "Birim Fiyat TL": format_tl(single_tl),
                "Toplam Fiyat EUR": format_eur(part_total_eur),
                "Toplam Fiyat TL": format_tl(part_total_tl),
            })

        st.dataframe(
            pd.DataFrame(list_rows),
            use_container_width=True,
            hide_index=True,
        )

        st.divider()
        st.markdown("### Teklif Bilgileri")

        with st.container(border=True):
            (
                quote_col1,
                quote_col2,
                quote_col3,
                quote_col4,
            ) = st.columns([1.2, 2, 1.5, 1.5])

            with quote_col1:
                profit_rate_percent = st.number_input(
                    "Kâr oranı (%)",
                    min_value=0.0,
                    value=30.0,
                    step=1.0,
                    format="%.2f",
                    key="quote_profit_rate",
                )

            with quote_col2:
                prepared_by = st.text_input(
                    "Teklif hazırlayan",
                    value="Erkan Engin",
                    key="quote_prepared_by",
                )

            with quote_col3:
                quote_date = st.date_input(
                    "Teklif Tarihi",
                    value=datetime.now().date(),
                    format="DD.MM.YYYY",
                    key="quote_date",
                )

            with quote_col4:
                quote_number = st.text_input(
                    "Teklif Numarası",
                    placeholder="Örn. ITS-2026-001",
                    key="quote_number",
                )

            email_address = st.text_input(
                "E-posta",
                value="erkan.engin@itssystems.com.tr",
                disabled=True,
                key="quote_email",
            )

        profit_multiplier = (
            float(profit_rate_percent) / 100
        )
        profit_eur = (
            list_total_cost_eur * profit_multiplier
        )
        profit_tl = (
            list_total_cost_tl * profit_multiplier
        )
        quote_total_eur = (
            list_total_cost_eur + profit_eur
        )
        quote_total_tl = (
            list_total_cost_tl + profit_tl
        )

        st.markdown("### Teklif Özeti")

        (
            summary_cost_col,
            summary_profit_col,
            summary_total_col,
        ) = st.columns(3)

        with summary_cost_col:
            st.metric(
                "Toplam Maliyet",
                (
                    f"{format_eur(list_total_cost_eur)} / "
                    f"{format_tl(list_total_cost_tl)}"
                ),
            )

        with summary_profit_col:
            st.metric(
                f"Kâr (%{format_number(profit_rate_percent, 2)})",
                (
                    f"{format_eur(profit_eur)} / "
                    f"{format_tl(profit_tl)}"
                ),
            )

        with summary_total_col:
            st.metric(
                "Maliyet + Kâr",
                (
                    f"{format_eur(quote_total_eur)} / "
                    f"{format_tl(quote_total_tl)}"
                ),
            )

        quote_ready = bool(
            prepared_by.strip()
            and quote_number.strip()
        )

        if not quote_ready:
            st.info(
                "Teklif dosyasını indirebilmek için Teklif Hazırlayan "
                "ve Teklif Numarası alanlarını doldur."
            )

        safe_quote_number = "".join(
            character
            if character.isalnum()
            or character in {"-", "_"}
            else "_"
            for character in quote_number.strip()
        ) or "Teklif"

        download_col, delete_col = st.columns([1, 2])

        with download_col:
            st.download_button(
                "Teklifi indir (.xlsx)",
                data=build_quote_xlsx(
                    exchange_rate,
                    profit_rate_percent,
                    prepared_by.strip(),
                    quote_date,
                    quote_number.strip(),
                    email_address,
                ),
                file_name=(
                    f"ITSSystems_{safe_quote_number}_"
                    f"{quote_date:%Y-%m-%d}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                type="primary",
                key="download_quote_xlsx",
                use_container_width=True,
                disabled=not quote_ready,
            )

        @st.dialog("Tüm listeyi sil")
        def confirm_delete_all_parts():
            st.warning(
                "Listedeki tüm parçalar ve bu parçalara bağlı maliyet "
                "kayıtları kalıcı olarak silinecek. Emin misin?"
            )

            cancel_col, confirm_col = st.columns(2)

            with cancel_col:
                if st.button(
                    "Vazgeç",
                    use_container_width=True,
                    key="cancel_delete_all_parts",
                ):
                    st.rerun()

            with confirm_col:
                if st.button(
                    "Evet, tamamını sil",
                    type="primary",
                    use_container_width=True,
                    key="confirm_delete_all_parts",
                ):
                    try:
                        deleted_count = len(parts)

                        for part in parts:
                            db.table("parcalar").delete().eq(
                                "id",
                                part["id"],
                            ).execute()

                        st.session_state[
                            "deleted_all_parts_count"
                        ] = deleted_count
                        st.rerun()
                    except Exception as error:
                        st.error(
                            "Liste silinirken bir hata oluştu. "
                            f"Detay: {error}"
                        )

        with delete_col:
            if st.button(
                "Tüm Listeyi Sil",
                type="primary",
                use_container_width=True,
                key="open_delete_all_parts_dialog",
            ):
                confirm_delete_all_parts()


# =========================================================
# TEDARİKÇİ LİSTESİ
# =========================================================
with tab_suppliers:
    st.subheader("Tedarikçi Listesi")

    with st.expander(
        "Yeni tedarikçi kaydı",
        expanded=False,
    ):
        with st.form(
            "new_supplier_material_form",
            clear_on_submit=True,
        ):
            row1_col1, row1_col2, row1_col3 = st.columns(
                [2.5, 1.2, 1]
            )

            with row1_col1:
                supplier_material_name = st.text_input(
                    "Malzeme adı"
                )

            with row1_col2:
                supplier_unit_price_text = st.text_input(
                    "Birim fiyat",
                    value="0,00",
                )

            with row1_col3:
                supplier_currency = st.selectbox(
                    "Para birimi",
                    ["EUR", "TL"],
                )

            row2_col1, row2_col2 = st.columns(2)

            with row2_col1:
                supplier_name = st.text_input(
                    "Tedarikçi adı"
                )

            with row2_col2:
                supplier_web_link = st.text_input(
                    "Tedarikçi web linki",
                    placeholder="https://firma.com",
                )

            supplier_description = st.text_area(
                "Açıklama",
                placeholder=(
                    "Malzeme, stok, termin veya diğer notlar"
                ),
                height=100,
            )

            save_supplier = st.form_submit_button(
                "Tedarikçi kaydını ekle",
                type="primary",
                use_container_width=True,
            )

            if save_supplier:
                supplier_unit_price = parse_decimal(
                    supplier_unit_price_text
                )

                if not supplier_material_name.strip():
                    st.error("Malzeme adı boş bırakılamaz.")
                elif not supplier_name.strip():
                    st.error("Tedarikçi adı boş bırakılamaz.")
                elif (
                    supplier_unit_price is None
                    or supplier_unit_price < 0
                ):
                    st.error("Geçerli bir birim fiyat gir.")
                else:
                    db.table(
                        "tedarikci_malzemeleri"
                    ).insert(
                        {
                            "malzeme_adi": (
                                supplier_material_name.strip()
                            ),
                            "birim_fiyat": (
                                supplier_unit_price
                            ),
                            "para_birimi": (
                                supplier_currency
                            ),
                            "tedarikci_adi": (
                                supplier_name.strip()
                            ),
                            "web_linki": normalize_web_url(
                                supplier_web_link
                            ),
                            "aciklama": (
                                supplier_description.strip()
                            ),
                        }
                    ).execute()

                    st.success(
                        "Tedarikçi kaydı eklendi."
                    )
                    st.rerun()

    st.divider()
    st.markdown("### Kayıtlı tedarikçi malzemeleri")

    transferred_material_name = st.session_state.pop(
        "supplier_material_transferred",
        None,
    )
    if transferred_material_name:
        st.success(
            f'{transferred_material_name} Fiyat Tanımları bölümüne '
            "malzeme olarak aktarıldı. Yoğunluk ve açıklama "
            "bilgilerini oradan tamamlayabilirsin."
        )

    supplier_materials = get_supplier_materials()

    if not supplier_materials:
        st.info("Henüz tedarikçi kaydı bulunmuyor.")
    else:
        existing_material_names = {
            item["ad"].strip().casefold()
            for item in prices
            if item["kategori"] == "Malzeme"
        }

        (
            header_material,
            header_price,
            header_currency,
            header_supplier,
            header_web,
            header_description,
            header_action,
        ) = st.columns(
            [2.2, 1.0, 0.7, 1.4, 2.2, 2.0, 1.45]
        )

        with header_material:
            st.markdown("**Malzeme Adı**")
        with header_price:
            st.markdown("**Birim Fiyat**")
        with header_currency:
            st.markdown("**Para Birimi**")
        with header_supplier:
            st.markdown("**Tedarikçi Adı**")
        with header_web:
            st.markdown("**Web Linki**")
        with header_description:
            st.markdown("**Açıklama**")
        with header_action:
            st.markdown("**İşlem**")

        st.divider()

        for item in supplier_materials:
            (
                col_material,
                col_price,
                col_currency,
                col_supplier,
                col_web,
                col_description,
                col_action,
            ) = st.columns(
                [2.2, 1.0, 0.7, 1.4, 2.2, 2.0, 1.45],
                vertical_alignment="center",
            )

            with col_material:
                st.write(item["malzeme_adi"])

            with col_price:
                st.write(
                    format_number(
                        float(item["birim_fiyat"]),
                        4,
                    )
                )

            with col_currency:
                st.write(item["para_birimi"])

            with col_supplier:
                st.write(item["tedarikci_adi"])

            with col_web:
                web_link = item.get("web_linki", "") or ""
                if web_link:
                    st.markdown(
                        f"[{web_link}]({web_link})"
                    )
                else:
                    st.write("—")

            with col_description:
                st.write(item.get("aciklama", "") or "—")

            material_key = (
                item["malzeme_adi"].strip().casefold()
            )
            already_transferred = (
                material_key in existing_material_names
            )

            with col_action:
                if already_transferred:
                    st.button(
                        "Aktarıldı",
                        key=(
                            f'supplier_already_transferred_'
                            f'{item["id"]}'
                        ),
                        disabled=True,
                        use_container_width=True,
                    )
                elif st.button(
                    "Fiyat Tanımlarına Aktar",
                    key=(
                        f'transfer_supplier_material_'
                        f'{item["id"]}'
                    ),
                    type="primary",
                    use_container_width=True,
                ):
                    source_currency = item["para_birimi"]
                    source_price = float(
                        item["birim_fiyat"]
                    )
                    eur_snapshot, _ = convert_price(
                        source_price,
                        source_currency,
                        exchange_rate,
                    )

                    db.table("fiyat_tanimlari").insert(
                        {
                            "kategori": "Malzeme",
                            "ad": item["malzeme_adi"].strip(),
                            "aciklama": "",
                            "kaynak_para_birimi": (
                                source_currency
                            ),
                            "kaynak_birim_fiyat": (
                                source_price
                            ),
                            "birim_fiyat_eur": (
                                eur_snapshot
                            ),
                            "yogunluk_g_cm3": None,
                        }
                    ).execute()

                    st.session_state[
                        "supplier_material_transferred"
                    ] = item["malzeme_adi"]
                    st.rerun()

            st.divider()

    with st.container(
        key="unavailable_materials_area",
        border=False,
    ):
        with st.expander(
            "Türkiye’de Bulunamayan Malzemeler",
            expanded=False,
        ):
            st.caption(
                "Bu alan yalnızca Türkiye’de tedarikçisi "
                "bulunamayan malzemeleri kaydetmek ve "
                "aramak içindir."
            )

            with st.form(
                "new_unavailable_material_form",
                clear_on_submit=True,
            ):
                unavailable_name = st.text_input(
                    "Bulunamayan malzeme adı"
                )

                add_unavailable = st.form_submit_button(
                    "Malzemeyi kırmızı listeye ekle",
                    use_container_width=True,
                )

                if add_unavailable:
                    cleaned_name = unavailable_name.strip()

                    if not cleaned_name:
                        st.error(
                            "Malzeme adı boş bırakılamaz."
                        )
                    else:
                        existing_unavailable = (
                            get_unavailable_materials()
                        )
                        exact_exists = any(
                            item[
                                "malzeme_adi"
                            ].strip().casefold()
                            == cleaned_name.casefold()
                            for item
                            in existing_unavailable
                        )

                        if exact_exists:
                            st.warning(
                                "Bu malzeme kırmızı "
                                "listede zaten kayıtlı."
                            )
                        else:
                            db.table(
                                "bulunamayan_malzemeler"
                            ).insert(
                                {
                                    "malzeme_adi": (
                                        cleaned_name
                                    ),
                                }
                            ).execute()

                            st.success(
                                "Bulunamayan malzeme "
                                "kaydedildi."
                            )
                            st.rerun()

            unavailable_materials = (
                get_unavailable_materials()
            )

            unavailable_search = st.text_input(
                "Bulunamayan malzemelerde ara",
                placeholder=(
                    "Örn. Stainless Steel 15-5PH H900 "
                    "veya 15-5PH"
                ),
                key="unavailable_material_search",
            )

            if unavailable_search.strip():
                search_value = (
                    unavailable_search.strip().casefold()
                )
                filtered_unavailable = [
                    item
                    for item in unavailable_materials
                    if search_value
                    in item["malzeme_adi"].casefold()
                ]
            else:
                filtered_unavailable = (
                    unavailable_materials
                )

            if not filtered_unavailable:
                if unavailable_search.strip():
                    st.warning(
                        "Aramayla eşleşen bulunamayan "
                        "malzeme yok."
                    )
                else:
                    st.info(
                        "Henüz bulunamayan malzeme kaydı yok."
                    )
            else:
                unavailable_rows = [
                    {
                        "Malzeme Adı": item["malzeme_adi"],
                    }
                    for item in filtered_unavailable
                ]

                st.dataframe(
                    pd.DataFrame(unavailable_rows),
                    use_container_width=True,
                    hide_index=True,
                )

